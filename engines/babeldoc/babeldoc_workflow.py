import sys
import io
import os
import yaml
import subprocess
import csv
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# Windows 控制台 UTF-8 编码支持
if sys.platform == 'win32':
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    if isinstance(sys.stdout, io.TextIOWrapper):
        sys.stdout.reconfigure(encoding='utf-8')
    if isinstance(sys.stderr, io.TextIOWrapper):
        sys.stderr.reconfigure(encoding='utf-8')

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# ==================== 配置管理 ====================

class Config:
    """配置管理类 - 加载和管理 config.yaml 配置文件"""

    def __init__(self, config_path="config.yaml"):
        """
        初始化配置

        Args:
            config_path: 配置文件路径，默认为 config.yaml
        """
        with open(config_path, 'r', encoding='utf-8') as f:
            self.data = yaml.safe_load(f)

    @property
    def api_key(self):
        """获取 API 密钥"""
        return self.data['api']['translation_api_key']

    @property
    def api_base_url(self):
        """获取 API 基础 URL"""
        return self.data['api']['translation_api_base_url']

    @property
    def api_model(self):
        """获取 API 模型名称"""
        return self.data['api']['translation_api_model']

    @property
    def input_dir(self):
        """获取输入目录路径"""
        return Path(self.data['paths']['input_base'])

    @property
    def output_dir(self):
        """获取输出目录路径"""
        return Path(self.data['paths']['output_base'])

    @property
    def terminology_dir(self):
        """获取术语库目录路径"""
        return Path(self.data['paths']['terminology_folder'])

    @property
    def pdf_modes(self):
        """获取 PDF 输出模式列表"""
        return self.data['babeldoc']['pdf_modes']

    @property
    def bilingual_settings(self):
        """获取双语 PDF 配置"""
        return self.data['babeldoc']['bilingual_settings']

    @property
    def qps(self):
        """获取翻译并发数（每秒请求数）"""
        return self.data['babeldoc']['qps']

    @property
    def skip_scanned_detection(self):
        """是否跳过扫描 PDF 检测"""
        return self.data['babeldoc'].get('skip_scanned_detection', True)

    @property
    def max_concurrent_files(self):
        """获取最大并发文件数"""
        return self.data['batch']['max_concurrent_files']

    @property
    def resume_enabled(self):
        """是否启用断点续传"""
        return self.data['batch']['resume_enabled']


# ==================== 术语库管理 ====================

class GlossaryManager:
    """术语库管理类 - 加载和处理专业术语翻译"""

    def __init__(self, terminology_dir):
        """
        初始化术语库管理器

        Args:
            terminology_dir: 术语库目录路径
        """
        self.terminology_dir = Path(terminology_dir)

    def load_from_excel(self):
        """
        从 Excel 文件加载术语库

        Returns:
            dict: 术语字典 {英文: 中文}
        """
        if not load_workbook:
            print("警告: openpyxl 未安装，无法加载术语库")
            return {}

        glossary = {}
        excel_files = list(self.terminology_dir.glob("*.xlsx")) + \
                     list(self.terminology_dir.glob("*.xls"))

        for excel_file in excel_files:
            try:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    if ws.max_row <= 1:
                        continue

                    # 从第 2 行开始读取（跳过标题）
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[0] and row[1]:
                            glossary[str(row[0]).strip()] = str(row[1]).strip()

                workbook.close()
            except Exception as e:
                print(f"加载术语库失败: {excel_file} - {e}")

        return glossary

    def export_to_csv(self, glossary, output_path):
        """
        将术语库导出为 BabelDOC 格式的 CSV

        Args:
            glossary: 术语字典
            output_path: 输出 CSV 文件路径

        Returns:
            str: CSV 文件路径，如果术语库为空则返回 None
        """
        if not glossary:
            return None

        with open(output_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['source', 'target', 'tgt_lng'])
            for en, zh in glossary.items():
                writer.writerow([en, zh, 'zh-CN'])

        return str(output_path)


# ==================== BabelDOC 命令构建器 ====================

class BabelDOCCommandBuilder:
    """BabelDOC 命令构建器 - 生成 BabelDOC 命令行参数"""

    # BabelDOC 可执行文件路径
    BABELDOC_EXE = r"C:\ProgramData\anaconda3\envs\babeldoc\Scripts\babeldoc.exe"

    def __init__(self, config):
        """
        初始化命令构建器

        Args:
            config: Config 对象
        """
        self.config = config

    def build(self, pdf_path, glossary_csv=None):
        """
        构建 BabelDOC 命令

        Args:
            pdf_path: PDF 文件路径
            glossary_csv: 术语库 CSV 文件路径（可选）

        Returns:
            list: 命令参数列表
        """
        cmd = [
            self.BABELDOC_EXE,
            "--files", str(pdf_path),
            "--output", str(self.config.output_dir),

            # API 配置
            "--openai",
            "--openai-model", self.config.api_model,
            "--openai-base-url", self.config.api_base_url,
            "--openai-api-key", self.config.api_key,

            # 性能配置
            "--qps", str(self.config.qps),
        ]

        # 跳过扫描检测
        if self.config.skip_scanned_detection:
            cmd.append("--skip-scanned-detection")

        # 添加术语库
        if glossary_csv:
            cmd.extend(["--glossary-files", glossary_csv])

        # PDF 输出模式
        pdf_modes = self.config.pdf_modes
        if 'translated_only' in pdf_modes and 'bilingual' not in pdf_modes:
            cmd.append("--no-dual")  # 只生成译文 PDF
        elif 'bilingual' in pdf_modes and 'translated_only' not in pdf_modes:
            cmd.append("--no-mono")  # 只生成双语 PDF

        # 双语 PDF 配置
        if 'bilingual' in pdf_modes:
            settings = self.config.bilingual_settings
            if settings.get('translated_first', True):
                cmd.append("--dual-translate-first")
            if settings.get('alternating_pages', False):
                cmd.append("--use-alternating-pages-dual")
            if not settings.get('watermark', False):
                cmd.append("--no-watermark")

        return cmd


# ==================== PDF 处理器 ====================

class PDFProcessor:
    """PDF 处理器 - 处理单个 PDF 文件的翻译"""

    def __init__(self, config):
        """
        初始化 PDF 处理器

        Args:
            config: Config 对象
        """
        self.config = config
        self.glossary_manager = GlossaryManager(config.terminology_dir)
        self.command_builder = BabelDOCCommandBuilder(config)

        # 创建必要的目录
        self.config.input_dir.mkdir(exist_ok=True)
        self.config.output_dir.mkdir(exist_ok=True)
        self.config.terminology_dir.mkdir(exist_ok=True)

    def process(self, pdf_path):
        """
        处理单个 PDF 文件

        Args:
            pdf_path: PDF 文件路径

        Returns:
            bool: 处理成功返回 True，否则返回 False
        """
        print(f"\n{'='*60}")
        print(f"处理文件: {pdf_path}")
        print(f"{'='*60}\n")

        # 1. 加载术语库
        print("加载术语库...")
        glossary = self.glossary_manager.load_from_excel()
        glossary_csv = None

        if glossary:
            print(f"✓ 加载了 {len(glossary)} 条术语")
            csv_path = self.config.output_dir / "glossary.csv"
            glossary_csv = self.glossary_manager.export_to_csv(glossary, csv_path)

        # 2. 构建命令
        cmd = self.command_builder.build(pdf_path, glossary_csv)

        # 3. 执行翻译
        print("\n开始翻译...")
        print("提示: 翻译过程可能需要数分钟到数十分钟，请耐心等待\n")

        try:
            result = subprocess.run(
                cmd,
                check=True,
                timeout=3600,  # 1 小时超时
                text=True
            )
            print("\n✓ 翻译完成！")
            print(f"输出目录: {self.config.output_dir}")
            return True

        except subprocess.TimeoutExpired:
            print("\n✗ 处理超时（1小时）")
            return False

        except subprocess.CalledProcessError as e:
            print(f"\n✗ 处理失败")
            return False

        except FileNotFoundError:
            print("\n✗ BabelDOC 未找到")
            print("请检查环境配置")
            return False


# ==================== 批量处理器 ====================

class BatchProcessor:
    """批量处理器 - 批量处理多个 PDF 文件"""

    def __init__(self, config):
        """
        初始化批量处理器

        Args:
            config: Config 对象
        """
        self.config = config
        self.pdf_processor = PDFProcessor(config)

    def scan_pdf_files(self):
        """
        扫描输入目录中的所有 PDF 文件

        Returns:
            list: PDF 文件路径列表
        """
        pdf_files = list(self.config.input_dir.rglob("*.pdf"))

        # 过滤临时文件
        pdf_files = [
            f for f in pdf_files
            if not any([
                '_compressed.pdf' in f.name,
                '_part' in f.name,
                'temp_splits' in str(f)
            ])
        ]

        return pdf_files

    def is_completed(self, pdf_path):
        """
        检查文件是否已处理完成（用于断点续传）

        Args:
            pdf_path: PDF 文件路径

        Returns:
            bool: 已完成返回 True，否则返回 False
        """
        pdf_name = pdf_path.stem
        output_dir = self.config.output_dir
        pdf_modes = self.config.pdf_modes

        expected_files = []
        if 'translated_only' in pdf_modes:
            expected_files.append(output_dir / f"{pdf_name}_translated.pdf")
        if 'bilingual' in pdf_modes:
            expected_files.append(output_dir / f"{pdf_name}_dual.pdf")

        return expected_files and all(f.exists() for f in expected_files)

    def process(self):
        """
        批量处理所有 PDF 文件

        Returns:
            tuple: (成功数, 失败数)
        """
        print("\n" + "="*60)
        print("BabelDOC 批量处理模式")
        print("="*60 + "\n")

        # 1. 扫描文件
        pdf_files = self.scan_pdf_files()
        if not pdf_files:
            print("✗ input/ 目录中没有找到 PDF 文件")
            return 0, 0

        print(f"找到 {len(pdf_files)} 个PDF文件\n")

        # 2. 断点续传过滤
        if self.config.resume_enabled:
            to_process = []
            skipped = []

            for pdf in pdf_files:
                if self.is_completed(pdf):
                    skipped.append(pdf)
                else:
                    to_process.append(pdf)

            if skipped:
                print(f"✓ 跳过 {len(skipped)} 个已完成的文件")

            pdf_files = to_process

        if not pdf_files:
            print("\n✓ 所有文件都已处理完成！")
            return 0, 0

        # 3. 显示待处理文件
        print(f"\n待处理: {len(pdf_files)} 个文件")
        for i, pdf in enumerate(pdf_files, 1):
            print(f"  {i}. {pdf.name}")

        # 4. 用户确认
        confirm = input(f"\n确认开始批量处理？[y/N]: ").strip().lower()
        if confirm != 'y':
            print("已取消")
            return 0, 0

        # 5. 并发处理
        max_workers = self.config.max_concurrent_files
        print(f"\n开始处理（并发数: {max_workers}）...\n")

        success_count = 0
        failure_count = 0

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_pdf = {
                executor.submit(self.pdf_processor.process, pdf): pdf
                for pdf in pdf_files
            }

            for future in as_completed(future_to_pdf):
                pdf = future_to_pdf[future]
                try:
                    success = future.result()
                    if success:
                        success_count += 1
                        print(f"✓ 完成: {pdf.name}")
                    else:
                        failure_count += 1
                        print(f"✗ 失败: {pdf.name}")
                except Exception as e:
                    failure_count += 1
                    print(f"✗ 异常: {pdf.name} - {e}")

        # 6. 汇总
        print("\n" + "="*60)
        print("批量处理完成！")
        print(f"  成功: {success_count} 个文件")
        print(f"  失败: {failure_count} 个文件")
        print(f"  总计: {len(pdf_files)} 个文件")
        print("="*60)

        return success_count, failure_count


# ==================== 命令行界面 ====================

def show_help():
    """显示帮助信息"""
    help_text = """
BabelDOC 翻译工作流 - 使用帮助

用法:
    python babeldoc_workflow.py [选项] [PDF文件路径]

选项:
    无参数          批量处理 input/ 目录中的所有 PDF 文件
    <PDF路径>       处理指定的单个 PDF 文件
    --help, -h     显示此帮助信息

示例:
    # 批量处理
    python babeldoc_workflow.py

    # 单文件处理
    python babeldoc_workflow.py input/paper.pdf

功能特性:
    ✓ 单文件和批量处理
    ✓ 自定义术语库支持
    ✓ 断点续传（跳过已完成文件）
    ✓ 高并发翻译（可配置）
    ✓ 双语/纯译文 PDF 输出

目录结构:
    input/          输入 PDF 文件
    output/         输出翻译结果
    terminology/    术语库 Excel 文件
    config.yaml     配置文件

配置文件:
    编辑 config.yaml 修改：
    - API 密钥和模型
    - 并发数和性能设置
    - PDF 输出模式
    - 双语 PDF 配置

更多信息请查看 README.md
"""
    print(help_text)


def main():
    """主函数 - 命令行入口"""
    # 显示帮助
    if len(sys.argv) > 1 and sys.argv[1] in ['--help', '-h']:
        show_help()
        sys.exit(0)

    try:
        # 加载配置
        config = Config()

        # 单文件处理模式
        if len(sys.argv) > 1:
            pdf_path = Path(sys.argv[1])

            if not pdf_path.exists():
                print(f"✗ 文件不存在: {pdf_path}")
                sys.exit(1)

            processor = PDFProcessor(config)
            success = processor.process(pdf_path)
            sys.exit(0 if success else 1)

        # 批量处理模式
        else:
            batch = BatchProcessor(config)
            success_count, failure_count = batch.process()
            sys.exit(0 if failure_count == 0 else 1)

    except FileNotFoundError:
        print("✗ 配置文件 config.yaml 未找到")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\n用户中断操作")
        sys.exit(130)
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
