"""
主流程脚本 - 重构版

核心职责：流程编排器 (Orchestrator)
- 协调各个模块完成文档翻译工作流
- 负责高层流程控制，不包含具体业务逻辑实现

重构记录（2025-12-15）：
1. 提取 MinerU 批处理逻辑 → mineru_batch_processor.py
2. 提取翻译任务管理逻辑 → translation_task_manager.py
3. 提取内容处理辅助函数 → content_helpers.py
4. 代码量：1822行 → 1110行（减少39%）
5. 平均方法长度：76行 → 约35行（减少54%）
"""

import yaml
import os
import sys
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed, ThreadPoolExecutor, wait, FIRST_COMPLETED
from jinja2 import Template
import shutil
import threading
import queue
import time
import hashlib

from mineru_client import MinerUClient, FileTask, TaskState
from mineru_parser import MinerUParser
from article_translator import ArticleTranslator
from logger import Logger
from format_converter import FormatConverter
from outline_generator import OutlineGenerator
from path_manager import PathManager
from resume_manager import ResumeManager
from mineru_batch_processor import MinerUBatchProcessor
from translation_task_manager import TranslationTaskManager
from content_helpers import (
    process_images, merge_split_texts,
    group_narrow_images, get_chapter_context
)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None


class DocumentProcessor:
    """文档处理主类"""

    def __init__(self, config_path="config.yaml"):
        """
        初始化文档处理器

        Args:
            config_path: 配置文件路径
        """
        # 加载配置
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)

        self.logger = Logger()
        self.output_base = Path(self.config['paths']['output_base'])

        # 初始化MinerU客户端
        self.mineru = MinerUClient(
            api_token=self.config['api']['mineru_token'],
            model_version="pipeline",
            verify_ssl=False,
            max_retries=self.config['retry']['mineru_max_retries']
        )

        # 初始化解析器（修改输出目录到output/MinerU）
        mineru_output_dir = self.output_base / self.config['output']['mineru_folder']
        self.parser = MinerUParser(output_dir=str(mineru_output_dir))

        # 初始化格式转换器
        self.converter = FormatConverter(self.config, self.logger, self.output_base)

        # 初始化大纲生成器
        self.outline_gen = OutlineGenerator(self.config, self.logger, self.output_base)

        # 初始化路径管理器
        self.path_mgr = PathManager(self.config, self.logger)

        # 初始化断点续传管理器
        self.resume_mgr = ResumeManager(self.logger)

        # 初始化文件夹结构
        self._init_directories()

    def _init_directories(self):
        """初始化所需的文件夹结构"""
        input_base = Path(self.config['paths']['input_base'])
        output_base = Path(self.config['paths']['output_base'])
        terminology_folder = Path(self.config['paths']['terminology_folder'])

        # 输出文件夹名称
        mineru_folder = self.config['output']['mineru_folder']
        html_folder = self.config['output']['html_folder']
        pdf_folder = self.config['output']['pdf_folder']
        docx_folder = self.config['output']['docx_folder']
        cache_folder = self.config['output']['cache_folder']

        # 创建所有必要的目录
        folders = [
            input_base,
            terminology_folder,
            output_base / mineru_folder,
            output_base / html_folder,
            output_base / pdf_folder,
            output_base / docx_folder,
            output_base / cache_folder / 'outlines',
        ]

        for folder in folders:
            folder.mkdir(parents=True, exist_ok=True)

        self.logger.info(f"文件夹结构初始化完成")

        # 清理残留的临时文件
        self._cleanup_temp_files()

    def _cleanup_temp_files(self):
        """清理残留的临时分割和合并文件"""
        input_base = Path(self.config['paths']['input_base'])
        output_base = Path(self.config['paths']['output_base'])
        mineru_folder = self.config['output']['mineru_folder']

        cleanup_count = 0

        # 1. 清理 input 目录下的 temp_splits
        for temp_dir in input_base.rglob("temp_splits"):
            if temp_dir.is_dir():
                try:
                    import shutil
                    shutil.rmtree(temp_dir)
                    cleanup_count += 1
                except Exception as e:
                    self.logger.warning(f"无法清理 {temp_dir}: {e}")

        # 2. 清理 output/MinerU 下的 temp_parts
        mineru_base = output_base / mineru_folder
        if mineru_base.exists():
            for temp_dir in mineru_base.rglob("temp_parts"):
                if temp_dir.is_dir():
                    try:
                        import shutil
                        shutil.rmtree(temp_dir)
                        cleanup_count += 1
                    except Exception as e:
                        self.logger.warning(f"无法清理 {temp_dir}: {e}")

        # 3. 清理 input 目录下的 _compressed.pdf 文件（旧压缩文件）
        for compressed_file in input_base.rglob("*_compressed.pdf"):
            try:
                compressed_file.unlink()
                cleanup_count += 1
            except Exception as e:
                self.logger.warning(f"无法删除 {compressed_file}: {e}")

        if cleanup_count > 0:
            self.logger.info(f"已清理 {cleanup_count} 个临时文件/目录")

    def load_terminology_from_excel(self) -> dict:
        """
        从 terminology 文件夹下的 Excel 文件加载术语库

        Returns:
            术语字典 {"English": "中文"}
        """
        terminology_folder = Path(self.config['paths']['terminology_folder'])

        if not terminology_folder.exists():
            self.logger.warning(f"术语库文件夹不存在: {terminology_folder}")
            return {}

        if not load_workbook:
            self.logger.warning("openpyxl 未安装，无法读取 Excel 术语库")
            return {}

        glossary = {}
        excel_files = list(terminology_folder.glob("*.xlsx")) + list(terminology_folder.glob("*.xls"))

        if not excel_files:
            self.logger.warning(f"术语库文件夹中没有 Excel 文件: {terminology_folder}")
            return {}

        self.logger.info(f"正在加载术语库，共 {len(excel_files)} 个 Excel 文件...")

        for excel_file in excel_files:
            try:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    if sheet.max_row <= 1:
                        continue

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[0] and row[1]:
                            english_term = str(row[0]).strip()
                            chinese_term = str(row[1]).strip()

                            if english_term and chinese_term:
                                glossary[english_term] = chinese_term

                workbook.close()
                self.logger.info(f"  已加载: {excel_file.name} - {len(glossary)} 个术语")

            except Exception as e:
                self.logger.error(f"加载 Excel 文件失败: {excel_file.name} - {str(e)}")

        self.logger.success(f"术语库加载完成，共 {len(glossary)} 个术语")
        return glossary

    def batch_process(self):
        """
        批量处理 input 文件夹中的所有 PDF 文件
        流水线模式：MinerU 解析完一个文件立即开始翻译，不等待整批完成
        """
        self.logger.info("=" * 60)
        self.logger.info("批量处理模式（流水线优化 - 解析翻译并行）")
        self.logger.info("=" * 60)

        # 1. 扫描输入文件
        file_list = self.path_mgr.scan_input_files()

        if not file_list:
            self.logger.error("没有找到要处理的 PDF 文件")
            return

        # 2. 加载全局术语库（从 Excel）
        excel_glossary = self.load_terminology_from_excel()

        # 3. 使用断点续传管理器检查文件状态
        categorized = self.resume_mgr.categorize_files(file_list, self.path_mgr)

        # 如果全部已完成，直接返回
        if self.resume_mgr.is_all_completed(categorized):
            self.logger.success("所有文件已处理完成！")
            return []

        # 准备处理列表
        files_to_upload, ready_to_translate = self.resume_mgr.prepare_processing_lists(categorized)
        already_completed = [status.relative_path for status in categorized['completed']]

        # 4. 创建任务队列和结果收集
        translation_queue = queue.Queue()  # MinerU完成的文件放入此队列
        failed_files = []  # 记录MinerU失败的文件
        failed_files_lock = threading.Lock()
        results = []
        results_lock = threading.Lock()

        # 先把已有结果的文件加入队列
        for item in ready_to_translate:
            translation_queue.put(item)

        # 5. 启动翻译工作线程池
        max_workers = self.config['concurrency']['max_files']
        self.logger.info(f"\n>>> 启动翻译工作池 (并发数: {max_workers})...")

        stop_event = threading.Event()
        translation_futures = []

        executor = ProcessPoolExecutor(max_workers=max_workers)

        # 6. 如果有需要上传的文件，启动 MinerU 监控线程
        monitor_thread = None
        if files_to_upload:
            self.logger.info(f"\n>>> 启动 MinerU 上传和监控...")
            # 创建 MinerU 批处理器
            mineru_processor = MinerUBatchProcessor(
                self.mineru, self.logger, self.config, self.path_mgr
            )
            monitor_thread = threading.Thread(
                target=mineru_processor.upload_and_monitor,
                args=(files_to_upload, translation_queue, stop_event, failed_files, failed_files_lock),
                daemon=True
            )
            monitor_thread.start()
        else:
            # 没有需要上传的文件，直接设置停止事件
            stop_event.set()

        # 7. 翻译任务提交线程：从队列中取任务并提交到进程池
        # 只处理需要处理的文件（排除已完成的）
        total_files = len(ready_to_translate) + len(files_to_upload)

        if total_files == 0:
            self.logger.success("所有文件已处理完成！")
            return []

        self.logger.info(f"\n>>> 启动翻译任务调度 (共 {total_files} 个文件，已跳过 {len(already_completed)} 个)...")

        future_to_file = {}  # {future: relative_path}
        future_lock = threading.Lock()
        submitted_count = 0

        def submit_tasks():
            """从队列中取任务并提交到进程池"""
            nonlocal submitted_count
            while submitted_count < total_files:
                try:
                    # 从队列获取任务
                    relative_path, pdf_path, mineru_zip_path = translation_queue.get(timeout=5)

                    self.logger.info(f"[提交] {relative_path}")

                    # 提交翻译任务
                    future = executor.submit(
                        self._process_translation_only,
                        relative_path,
                        pdf_path,
                        excel_glossary,
                        mineru_zip_path
                    )

                    with future_lock:
                        future_to_file[future] = relative_path
                        submitted_count += 1

                except queue.Empty:
                    # 队列暂时为空
                    if stop_event.is_set() and translation_queue.empty():
                        # MinerU已完成且队列为空
                        break
                    continue

        # 启动提交线程
        submit_thread = threading.Thread(target=submit_tasks, daemon=True)
        submit_thread.start()

        # 8. 收集翻译结果
        self.logger.info(f"\n>>> 开始收集翻译结果...")
        success_count = 0
        failure_count = 0
        processed_count = 0

        if tqdm:
            pbar = tqdm(total=total_files, desc="总进度")

        # 使用 as_completed 收集结果
        while processed_count < total_files:
            # 等待至少有一个任务提交
            should_wait = False
            with future_lock:
                if not future_to_file:
                    # 检查是否还有失败的文件没统计
                    with failed_files_lock:
                        failed_count = len(failed_files)

                    # 如果已处理数 + 失败数 = 总数，说明全部完成
                    if processed_count + failed_count >= total_files:
                        break

                    # 如果没有任务但已经全部提交完，退出
                    if submitted_count + failed_count >= total_files:
                        break
                    # 否则需要等待
                    should_wait = True

            if should_wait:
                time.sleep(0.5)
                continue

            # 收集已完成的任务
            done_futures = []
            with future_lock:
                for future in list(future_to_file.keys()):
                    if future.done():
                        done_futures.append(future)

            for future in done_futures:
                with future_lock:
                    relative_path = future_to_file.pop(future)

                try:
                    result = future.result()
                    if result['success']:
                        success_count += 1
                        self.logger.success(f"✓ 完成: {relative_path}")
                    else:
                        failure_count += 1
                        self.logger.error(f"✗ 失败: {relative_path} - {result.get('error', 'Unknown')}")
                    results.append(result)

                except Exception as e:
                    failure_count += 1
                    self.logger.error(f"✗ 异常: {relative_path} - {str(e)}")
                    results.append({'success': False, 'file': relative_path, 'error': str(e)})

                processed_count += 1
                if tqdm:
                    pbar.update(1)

            # 检查是否全部完成
            if processed_count >= total_files:
                break

            # 短暂休眠
            if not done_futures:
                time.sleep(0.1)

        if tqdm:
            pbar.close()

        # 9. 等待所有线程结束
        submit_thread.join(timeout=10)
        if monitor_thread:
            monitor_thread.join(timeout=10)
        executor.shutdown(wait=True)

        # 10. 输出汇总（包含跳过的文件和失败的文件）
        # 将已完成的文件加入结果
        for relative_path in already_completed:
            results.append({
                'success': True,
                'file': relative_path,
                'skipped': True,
                'reason': '已存在完整输出'
            })

        # 将MinerU失败的文件加入结果
        with failed_files_lock:
            for relative_path, error_msg in failed_files:
                results.append({
                    'success': False,
                    'file': relative_path,
                    'error': error_msg
                })
                failure_count += 1

        total_count = len(results) + len(already_completed)
        self.logger.info("=" * 60)
        self.logger.info(f"批量处理完成！")
        self.logger.info(f"  成功: {success_count} 个文件")
        self.logger.info(f"  失败: {failure_count} 个文件")
        self.logger.info(f"  跳过: {len(already_completed)} 个文件（已完成）")
        self.logger.info(f"  总计: {len(file_list)} 个文件")
        self.logger.info("=" * 60)
        self.logger.info("翻译请求日志已保存到 logs/translation/ 目录")

        return results

    def _batch_upload_to_mineru(self, file_list: list) -> dict:
        """
        批量上传文件到 MinerU（每批最多200个文件）

        Args:
            file_list: [(relative_path, pdf_path), ...] 文件列表

        Returns:
            {relative_path: zip_path} 字典，映射相对路径到下载的ZIP文件路径
        """
        # 1. 检查哪些文件已经有结果，哪些需要上传
        files_to_upload = []
        existing_results = {}

        mineru_folder = self.config['output']['mineru_folder']
        mineru_dir = self.output_base / mineru_folder

        for relative_path, pdf_path in file_list:
            output_paths = self.path_mgr.get_output_paths(relative_path)
            expected_zip = Path(output_paths['mineru'])

            if expected_zip.exists():
                self.logger.info(f"✓ 已存在: {relative_path}")
                existing_results[relative_path] = str(expected_zip)
            else:
                files_to_upload.append((relative_path, pdf_path, output_paths))

        if not files_to_upload:
            self.logger.success("所有文件都已有 MinerU 解析结果，跳过上传")
            return existing_results

        self.logger.info(f"需要上传: {len(files_to_upload)} 个文件")
        self.logger.info(f"已有结果: {len(existing_results)} 个文件")

        # 2. 分批上传（每批200个）
        BATCH_SIZE = 200
        all_results = existing_results.copy()

        for batch_idx in range(0, len(files_to_upload), BATCH_SIZE):
            batch_files = files_to_upload[batch_idx:batch_idx + BATCH_SIZE]
            batch_num = batch_idx // BATCH_SIZE + 1
            total_batches = (len(files_to_upload) + BATCH_SIZE - 1) // BATCH_SIZE

            self.logger.info(f"\n--- 批次 {batch_num}/{total_batches} (共 {len(batch_files)} 个文件) ---")

            # 创建 FileTask 列表
            file_tasks = []
            for relative_path, pdf_path, output_paths in batch_files:
                file_task = FileTask(
                    file_name=Path(pdf_path).name,
                    file_path=pdf_path,
                    data_id=Path(pdf_path).stem
                )
                file_tasks.append(file_task)

            try:
                # 批量上传
                self.logger.info(f"正在上传批次 {batch_num}...")
                batch_id, _, split_info = self.mineru.batch_upload_files(file_tasks)

                # 等待完成
                self.logger.info(f"等待批次 {batch_num} 解析完成...")
                results = self.mineru.wait_for_completion(batch_id, poll_interval=10)

                # 下载结果
                self.logger.info(f"下载批次 {batch_num} 结果...")

                for i, (relative_path, pdf_path, output_paths) in enumerate(batch_files):
                    result = results[i]

                    if result.state == TaskState.DONE and result.full_zip_url:
                        # 下载到指定位置
                        expected_zip = Path(output_paths['mineru'])
                        expected_zip.parent.mkdir(parents=True, exist_ok=True)

                        # 生成保存文件名
                        base_name = Path(pdf_path).stem
                        zip_name = f"{base_name}_result.zip"

                        try:
                            zip_path = self.mineru.download_result(
                                result.full_zip_url,
                                str(expected_zip.parent),
                                zip_name
                            )

                            # 如果下载位置不是目标位置，移动文件
                            if Path(zip_path) != expected_zip:
                                shutil.move(zip_path, str(expected_zip))
                                zip_path = str(expected_zip)

                            all_results[relative_path] = zip_path
                            self.logger.success(f"✓ {relative_path}")

                        except Exception as e:
                            self.logger.error(f"✗ 下载失败: {relative_path} - {str(e)}")
                    else:
                        error_msg = result.err_msg or f"状态: {result.state.value}"
                        self.logger.error(f"✗ 解析失败: {relative_path} - {error_msg}")

            except Exception as e:
                self.logger.error(f"批次 {batch_num} 处理失败: {str(e)}")
                # 继续处理下一批
                continue

        self.logger.success(f"\nMinerU 批量上传完成！成功: {len(all_results)}/{len(file_list)} 个文件")
        return all_results

    def _process_translation_only(
        self,
        relative_path: str,
        pdf_path: str,
        excel_glossary: dict,
        mineru_zip_path: str
    ) -> dict:
        """
        只处理翻译和格式转换（MinerU 解析已完成）

        Args:
            relative_path: 相对路径
            pdf_path: PDF 绝对路径
            excel_glossary: Excel 术语库
            mineru_zip_path: MinerU 结果 ZIP 文件路径

        Returns:
            处理结果字典
        """
        try:
            if not mineru_zip_path or not Path(mineru_zip_path).exists():
                return {
                    'success': False,
                    'file': relative_path,
                    'error': 'MinerU 解析结果不存在'
                }

            output_paths = self.path_mgr.get_output_paths(relative_path)

            # 智能断点续传：检查 HTML 是否已存在
            html_original_path = Path(output_paths['html_original'])
            html_translated_path = Path(output_paths['html_translated'])

            if html_translated_path.exists() and html_original_path.exists():
                # HTML 已存在，直接加载
                self.logger.info(f"检测到已有 HTML，跳过翻译，只补全格式转换: {relative_path}")
                original_html = html_original_path.read_text(encoding='utf-8')
                translated_html = html_translated_path.read_text(encoding='utf-8')
            else:
                # HTML 不存在，需要完整处理
                # 1. 解析 MinerU 结果
                parsed = self.parser.parse_zip_result(
                    mineru_zip_path,
                    source_file_name=Path(pdf_path).name
                )

                extract_dir = self.parser.output_dir / Path(mineru_zip_path).stem

                # 2. 生成大纲
                outline = self.outline_gen.generate_outline(pdf_path, output_paths)

                # 3. 初始化翻译器
                translator = ArticleTranslator(
                    api_key=self.config['api']['translation_api_key'],
                    api_url=self.config['api']['translation_api_base_url'],
                    model=self.config['api']['translation_api_model'],
                    glossary=excel_glossary or {},
                    case_sensitive=False,
                    whole_word_only=True,
                    config=self.config
                )

                # 设置当前文件名（用于日志）
                translator.current_file = Path(relative_path).stem

                # 4. 处理内容并翻译
                original_html, translated_html = self.process_content(
                    parsed.json_content,
                    outline,
                    translator,
                    str(extract_dir),
                    output_paths
                )

            # 5. 导出格式（会智能跳过已存在的文件）
            self.converter.export_formats(original_html, translated_html, output_paths)

            return {
                'success': True,
                'file': relative_path,
                'output_paths': {k: str(v) for k, v in output_paths.items()}
            }

        except Exception as e:
            return {
                'success': False,
                'file': relative_path,
                'error': str(e)
            }

    def _process_single_file(self, relative_path: str, pdf_path: str, excel_glossary: dict) -> dict:
        """
        处理单个 PDF 文件（用于多进程调用）

        Args:
            relative_path: 相对路径
            pdf_path: PDF 绝对路径
            excel_glossary: Excel 术语库

        Returns:
            处理结果字典
        """
        try:
            output_paths = self.path_mgr.get_output_paths(relative_path)
            self.run(pdf_path, output_paths, excel_glossary)

            return {
                'success': True,
                'file': relative_path,
                'output_paths': {k: str(v) for k, v in output_paths.items()}
            }
        except Exception as e:
            return {
                'success': False,
                'file': relative_path,
                'error': str(e)
            }

    def run(self, pdf_path: str, output_paths: dict = None, excel_glossary: dict = None):
        """
        运行完整流程

        Args:
            pdf_path: PDF文件路径
            output_paths: 自定义输出路径字典（可选）
            excel_glossary: Excel术语库（可选）
        """
        self.logger.info("=" * 60)
        self.logger.info("开始处理文档")
        self.logger.info("=" * 60)

        try:
            # 步骤1: 生成大纲
            outline = self.outline_gen.generate_outline(pdf_path, output_paths)

            # 步骤2: MinerU解析
            content_list, extract_dir = self.parse_with_mineru(pdf_path, output_paths)

            # 步骤3: 使用 Excel 术语库（不使用 AI 生成的术语）
            combined_glossary = excel_glossary or {}
            
            if combined_glossary:
                self.logger.info(f"术语库加载完成: {len(combined_glossary)} 个术语")
            else:
                self.logger.warning("未找到术语库，将不进行术语预替换")

            # 步骤4: 初始化翻译器
            translator = ArticleTranslator(
                api_key=self.config['api']['translation_api_key'],
                api_url=self.config['api']['translation_api_base_url'],
                model=self.config['api']['translation_api_model'],
                glossary=combined_glossary,
                case_sensitive=False,
                whole_word_only=True,
                config=self.config
            )

            # 步骤5: 处理内容并翻译
            original_html, translated_html = self.process_content(
                content_list, outline, translator, extract_dir, output_paths
            )

            # 步骤6: 导出格式
            self.converter.export_formats(original_html, translated_html, output_paths)

            self.logger.info("=" * 60)
            self.logger.success("处理完成！")
            self.logger.info("=" * 60)

        except Exception as e:
            self.logger.error(f"处理失败: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def parse_with_mineru(self, pdf_path: str, output_paths: dict = None) -> tuple:
        """
        使用MinerU解析PDF

        Args:
            pdf_path: PDF文件路径
            output_paths: 自定义输出路径字典（可选）

        Returns:
            (content_list, extract_dir) - 内容列表和解压目录
        """
        self.logger.info("\n>>> 步骤2: 使用MinerU解析PDF...")

        # 确定ZIP保存路径（output/MinerU/相对路径）
        if output_paths and 'mineru' in output_paths:
            expected_zip = Path(output_paths['mineru'])
        else:
            mineru_folder = self.config['output']['mineru_folder']
            mineru_dir = self.output_base / mineru_folder
            pdf_name = Path(pdf_path).stem
            expected_zip = mineru_dir / f"{pdf_name}_result.zip"

        expected_zip.parent.mkdir(parents=True, exist_ok=True)

        # 检查是否已有解析结果
        if expected_zip.exists():
            self.logger.info("发现已有MinerU解析结果，直接加载...")
            parsed = self.parser.parse_zip_result(
                str(expected_zip),
                source_file_name=Path(pdf_path).name
            )
            # 获取解压目录
            extract_dir = self.parser.output_dir / Path(expected_zip).stem
            self.logger.success(f"解析结果已加载: {len(parsed.json_content)} 个内容块")
            return parsed.json_content, str(extract_dir)

        # 上传并解析
        # 使用文件路径的MD5哈希前16位作为 data_id（保证不超过128字符）
        data_id = hashlib.md5(pdf_path.encode('utf-8')).hexdigest()[:16]

        file_task = FileTask(
            file_name=Path(pdf_path).name,
            file_path=pdf_path,
            data_id=data_id
        )

        self.logger.info("正在上传PDF到MinerU...")
        batch_id, _, split_info = self.mineru.batch_upload_files([file_task])

        self.logger.info("等待MinerU解析完成...")
        results = self.mineru.wait_for_completion(batch_id, poll_interval=10)

        # 如果文件被分割了，需要合并结果
        if pdf_path in split_info:
            self.logger.info(f"文件被分割为 {len(split_info[pdf_path])} 部分，准备下载并合并...")

            # 下载所有部分
            part_zips = []
            page_offsets = []

            for part_idx, (task_idx, start_page, end_page) in enumerate(split_info[pdf_path], 1):
                result = results[task_idx]

                if result.state == TaskState.DONE and result.full_zip_url:
                    # 下载到临时目录
                    temp_dir = expected_zip.parent / "temp_parts"
                    temp_dir.mkdir(parents=True, exist_ok=True)

                    part_name = f"{Path(pdf_path).stem}_part{part_idx}_result.zip"
                    zip_path = self.mineru.download_result(
                        result.full_zip_url,
                        str(temp_dir),
                        part_name
                    )

                    part_zips.append(zip_path)
                    page_offsets.append(start_page)
                    self.logger.success(f"✓ Part {part_idx} 下载完成")
                else:
                    error_msg = result.err_msg or "未知错误"
                    raise RuntimeError(f"Part {part_idx} 解析失败: {error_msg}")

            # 合并所有部分
            self.logger.info("正在合并所有部分...")
            self.mineru._merge_mineru_results(part_zips, str(expected_zip), page_offsets)

            # 清理临时文件
            for part_zip in part_zips:
                try:
                    Path(part_zip).unlink()
                except:
                    pass

            self.logger.success("✓ 所有部分已合并")

        else:
            # 未分割：直接下载
            downloaded = self.mineru.download_all_results(results, str(expected_zip.parent))

            # 检查是否成功下载
            if not downloaded:
                error_msg = "MinerU解析失败，没有可下载的结果。"
                # 检查results中的失败原因
                for result in results:
                    if result.state == TaskState.FAILED:
                        reason = result.err_msg or '未知原因'
                        error_msg += f"\n失败原因: {reason}"
                        error_msg += "\n\n可能的解决方案:"
                        error_msg += "\n1. 检查PDF文件是否损坏或加密"
                        error_msg += "\n2. 尝试重新下载或转换PDF文件"
                        error_msg += "\n3. 检查PDF文件大小是否超过限制"
                self.logger.error(error_msg)
                raise RuntimeError(error_msg)

            # 获取下载的zip文件路径
            zip_path = list(downloaded.values())[0]

            # 如果下载位置不是目标位置，移动文件
            if Path(zip_path) != expected_zip:
                shutil.move(zip_path, str(expected_zip))

        # 解析ZIP
        parsed = self.parser.parse_zip_result(
            str(expected_zip),
            source_file_name=Path(pdf_path).name
        )

        # 获取解压目录
        extract_dir = self.parser.output_dir / Path(expected_zip).stem

        self.logger.success(f"解析完成: {len(parsed.json_content)} 个内容块")
        return parsed.json_content, str(extract_dir)

    def process_content(
        self,
        content_list: list,
        outline: dict,
        translator: ArticleTranslator,
        extract_dir: str,
        output_paths: dict = None
    ) -> tuple:
        """
        处理内容并翻译

        Args:
            content_list: MinerU返回的content_list
            outline: 文档大纲
            translator: 翻译器实例
            extract_dir: MinerU解压目录
            output_paths: 输出路径字典

        Returns:
            (original_html, translated_html) 元组
        """
        self.logger.info("\n>>> 步骤3: 处理内容并翻译...")

        # 按页分组
        pages = {}
        for item in content_list:
            page_idx = item.get('page_idx', 0)
            if page_idx not in pages:
                pages[page_idx] = []
            pages[page_idx].append(item)

        self.logger.info(f"共 {len(pages)} 页")

        # 处理图片：复制到HTML目录并更新路径
        process_images(content_list, extract_dir, output_paths, self.logger, self.config)

        # 极简合并：处理连字符断词和跨列分割
        for page_idx in pages.keys():
            items = pages[page_idx]
            merged_items = merge_split_texts(items)
            pages[page_idx] = merged_items

        # 创建翻译任务管理器
        task_mgr = TranslationTaskManager(self.logger, self.config)

        # 加载失败文本缓存
        failed_texts_cache = task_mgr.load_failed_cache()

        # 收集翻译任务
        tasks = task_mgr.collect_tasks(pages, outline, get_chapter_context)

        # 执行翻译
        translations = task_mgr.execute_translations(tasks, translator)

        # 分配翻译结果
        retry_stats = task_mgr.assign_results(tasks, translations, failed_texts_cache)

        # 更新失败日志
        task_mgr.update_failed_log(failed_texts_cache, retry_stats)

        # 生成HTML
        self.logger.info("正在生成HTML...")

        # 对图片进行智能分组（连续的窄长图片合并成一行）
        pages = group_narrow_images(pages, self.logger)

        original_html = self._render_html(pages, language='en')
        translated_html = self._render_html(pages, language='zh')

        self.logger.success("HTML已生成")

        return original_html, translated_html

    def _render_html(self, pages: dict, language: str) -> str:
        """渲染HTML"""
        with open('page_template.html', 'r', encoding='utf-8') as f:
            template = Template(f.read())

        return template.render(pages=pages, language=language)


def main():
    """命令行入口"""
    if len(sys.argv) == 1:
        interactive_mode()
        return

    if sys.argv[1] in ["--batch", "-b", "--interactive", "-i"]:
        interactive_mode()
    else:
        print(f"❌ 未知参数: {sys.argv[1]}")
        print("使用 'python main.py -h' 查看帮助")
        sys.exit(1)

def interactive_mode():
    """交互式命令行界面"""
    processor = DocumentProcessor()

    while True:
        print("\n" + "="*60)
        print("  MinerU 文档翻译工具 - 交互模式")
        print("="*60)
        print("\n请选择操作：")
        print("  [1] 批量处理（递归扫描 input/ 文件夹）")
        print("  [2] 查看配置信息")
        print("  [3] 查看输入文件列表")
        print("  [4] 清除缓存")
        print("  [0] 退出")
        print()

        choice = input("请输入选项 [0-4]: ").strip()

        if choice == "0":
            print("\n再见！")
            break
        elif choice == "1":
            batch_mode_interactive(processor)
        elif choice == "2":
            show_config(processor)
        elif choice == "3":
            show_input_files(processor)
        elif choice == "4":
            clear_cache(processor)
        else:
            print("❌ 无效选项，请重新选择")


def batch_mode_interactive(processor):
    """批量处理交互模式"""
    print("\n" + "-"*60)
    print("  批量处理模式")
    print("-"*60)

    file_list = processor.path_mgr.scan_input_files()

    if not file_list:
        print("\n❌ input/ 文件夹中没有找到 PDF 文件")
        print("   请先将 PDF 文件放入 input/ 文件夹")
        input("\n按回车键继续...")
        return

    print(f"\n找到 {len(file_list)} 个 PDF 文件:")
    for i, (rel_path, abs_path) in enumerate(file_list[:10], 1):
        print(f"  {i}. {rel_path}")

    if len(file_list) > 10:
        print(f"  ... 还有 {len(file_list) - 10} 个文件")

    print(f"\n并发配置:")
    print(f"  - 文件并发数: {processor.config['concurrency']['max_files']}")
    print(f"  - 翻译并发数: {processor.config['concurrency']['initial_translation_workers']} (初始)")

    confirm = input(f"\n确认开始批量处理？[y/N]: ").strip().lower()

    if confirm != 'y':
        print("已取消")
        return

    try:
        print("\n开始批量处理...")
        processor.batch_process()
        print("\n✓ 批量处理完成！")
    except Exception as e:
        print(f"\n❌ 批量处理失败: {str(e)}")

    input("\n按回车键继续...")


def show_config(processor):
    """显示配置信息"""
    print("\n" + "-"*60)
    print("  当前配置信息")
    print("-"*60)

    config = processor.config

    print("\n📡 API 配置:")
    print(f"  MinerU Token: {'已配置' if config['api']['mineru_token'] != 'YOUR_MINERU_TOKEN' else '❌ 未配置'}")
    print(f"  Translation API: {config['api']['translation_api_base_url']}")
    print(f"  Translation Model: {config['api']['translation_api_model']}")

    print("\n🔄 并发配置:")
    print(f"  文件并发数: {config['concurrency']['max_files']}")
    print(f"  翻译并发数: {config['concurrency']['initial_translation_workers']}-{config['concurrency']['max_translation_workers']}")

    input("\n按回车键继续...")


def show_input_files(processor):
    """显示输入文件列表"""
    print("\n" + "-"*60)
    print("  输入文件列表")
    print("-"*60)

    file_list = processor.path_mgr.scan_input_files()

    if not file_list:
        print("\n❌ input/ 文件夹中没有找到 PDF 文件")
    else:
        print(f"\n找到 {len(file_list)} 个 PDF 文件:\n")
        for i, (rel_path, abs_path) in enumerate(file_list, 1):
            file_size = Path(abs_path).stat().st_size / (1024 * 1024)
            print(f"  {i:3d}. {rel_path:50s} ({file_size:.1f} MB)")

    input("\n按回车键继续...")


def clear_cache(processor):
    """清除缓存"""
    print("\n" + "-"*60)
    print("  清除缓存")
    print("-"*60)

    cache_dir = processor.output_base / "cache"

    if not cache_dir.exists():
        print("\n没有缓存需要清除")
        input("\n按回车键继续...")
        return

    total_size = 0
    file_count = 0
    for file in cache_dir.rglob("*"):
        if file.is_file():
            total_size += file.stat().st_size
            file_count += 1

    print(f"\n缓存统计:")
    print(f"  文件数: {file_count}")
    print(f"  总大小: {total_size / (1024 * 1024):.1f} MB")

    confirm = input("\n确认清除所有缓存？[y/N]: ").strip().lower()

    if confirm != 'y':
        print("已取消")
        input("\n按回车键继续...")
        return

    try:
        shutil.rmtree(cache_dir)
        cache_dir.mkdir(parents=True, exist_ok=True)
        print("\n✓ 缓存已清除")
    except Exception as e:
        print(f"\n❌ 清除失败: {str(e)}")

    input("\n按回车键继续...")


if __name__ == "__main__":
    main()