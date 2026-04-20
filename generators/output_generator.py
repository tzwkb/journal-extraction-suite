
import os
import sys
import re
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from core.logger import get_logger
from config import UserConfig
from core.pdf_utils import remove_file_with_retry, is_file_locked

logger = get_logger("output_generator")

_is_frozen = getattr(sys, 'frozen', False)
_project_base = Path(sys.executable).parent if _is_frozen else Path(__file__).parent


def _generate_with_retry(generate_fn, html_path: Path, output_path: Path, file_type_label: str) -> str:
    import time
    max_retries = UserConfig.PDF_GENERATION_MAX_RETRIES
    last_error = None
    for attempt in range(max_retries):
        try:
            generate_fn(str(html_path), str(output_path))
            if not output_path.exists():
                raise RuntimeError(f"{file_type_label}文件生成失败（文件不存在）")
            file_size = output_path.stat().st_size / 1024
            logger.info(f"✅ {file_type_label}生成成功: {output_path} ({file_size:.1f} KB)")
            return str(output_path)
        except (PermissionError, OSError) as e:
            last_error = e
            if attempt < max_retries - 1:
                logger.warning(f"⚠️ {file_type_label}生成失败（尝试 {attempt + 1}/{max_retries}）: {e}")
                time.sleep(UserConfig.PDF_GENERATION_RETRY_DELAY)
            else:
                if "Permission denied" in str(e) or "PermissionError" in str(type(e)):
                    error_msg = (
                        f"无法写入{file_type_label}文件（权限被拒绝）: {output_path}\n"
                        f"可能原因:\n"
                        f"  1. 文件正在被其他程序使用\n"
                        f"  2. 文件或文件夹权限不足\n"
                        f"  3. 磁盘空间不足\n"
                        f"请检查并重试"
                    )
                    logger.error(f"❌ {error_msg}")
                    raise PermissionError(error_msg) from e
                else:
                    raise
    if last_error:
        raise last_error
    raise RuntimeError(f"{file_type_label}生成失败（未知错误）")


class PDFGenerator:
    
    def __init__(self):
        self._check_dependencies()
    
    def _get_browser_path(self) -> Optional[str]:
        import platform
        import glob

        if platform.system() == 'Windows':
            browser_search_paths = [
                _project_base / '_internal' / 'playwright_browsers' / 'chromium' / 'chrome-win' / 'chrome.exe',
                _project_base / '_internal' / 'playwright_browsers' / 'chromium_headless_shell' / 'chrome-win' / 'headless_shell.exe',
                _project_base / 'playwright_browsers' / 'chromium' / 'chrome-win' / 'chrome.exe',
                _project_base / 'playwright_browsers' / 'chromium_headless_shell' / 'chrome-win' / 'headless_shell.exe',
                _project_base / 'playwright_browsers' / 'chromium_downloaded' / 'chrome-win' / 'chrome.exe',
            ]

            for browser_path in browser_search_paths:
                if browser_path.exists():
                    logger.info(f"🔍 找到浏览器: {browser_path.relative_to(_project_base)}")
                    return str(browser_path)

            if not _is_frozen:
                local_app_data = os.environ.get('LOCALAPPDATA', '')
                if local_app_data:
                    ms_playwright_dir = Path(local_app_data) / 'ms-playwright'
                    if ms_playwright_dir.exists():
                        for pattern in [
                            ms_playwright_dir / 'chromium_headless_shell-*' / 'chrome-win' / 'headless_shell.exe',
                            ms_playwright_dir / 'chromium-*' / 'chrome-win' / 'chrome.exe',
                        ]:
                            matches = glob.glob(str(pattern))
                            if matches:
                                latest = sorted(matches)[-1]
                                logger.info(f"🔍 找到系统浏览器（备用）: {latest}")
                                return latest

        return None

    def _check_dependencies(self):
        try:
            from playwright.sync_api import sync_playwright

            browser_exe_path = self._get_browser_path()

            try:
                with sync_playwright() as p:
                    launch_options = {'headless': True}
                    if browser_exe_path:
                        launch_options['executable_path'] = browser_exe_path
                        logger.info(f"🔧 使用浏览器: {browser_exe_path}")
                    
                    browser = p.chromium.launch(**launch_options)
                    browser.close()
                    
                self._browser_exe_path = browser_exe_path if browser_exe_path else None
                logger.info(f"✅ 使用 Playwright 生成PDF（Chromium已就绪 - {'EXE模式' if _is_frozen else '开发模式'}）")
                return
                
            except Exception as browser_error:
                error_msg = str(browser_error)
                if "Executable doesn't exist" in error_msg or "browser is not installed" in error_msg.lower():
                    raise RuntimeError(
                        f"❌ Chromium浏览器未安装\n\n"
                        f"{'='*60}\n"
                        f"📥 解决方案：运行依赖管理脚本\n"
                        f"{'='*60}\n\n"
                        f"请在项目目录运行:\n"
                        f"  python setup_dependencies.py\n\n"
                        f"此脚本将自动:\n"
                        f"  ✓ 下载Chromium浏览器到项目目录\n"
                        f"  ✓ 配置所有必需的依赖\n"
                        f"  ✓ 确保开发和打包环境正常\n\n"
                        f"{'='*60}\n"
                        f"💡 或者手动安装:\n"
                        f"{'='*60}\n"
                        f"  1. 运行: playwright install chromium\n"
                        f"  2. 或从 https://commondatastorage.googleapis.com/chromium-browser-snapshots/index.html 下载\n"
                        f"  3. 解压到 playwright_browsers/chromium/ 目录"
                    )
                else:
                    raise
                
        except ImportError as e:
            raise RuntimeError(
                "❌ Playwright未安装\n"
                "请按以下步骤安装:\n"
                "  1. pip install playwright\n"
                "  2. playwright install chromium\n"
                f"详细错误: {e}"
            )
    
    def _generate_pdf(self, html_path: str, output_path: str):
        from playwright.sync_api import sync_playwright
        from pathlib import Path

        with sync_playwright() as p:
            launch_options = {
                'headless': True,
                'args': [
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--force-device-scale-factor=1',
                    '--disable-font-subpixel-positioning',
                    '--disable-lcd-text',
                    '--font-render-hinting=none',
                ]
            }
            if hasattr(self, '_browser_exe_path') and self._browser_exe_path:
                launch_options['executable_path'] = self._browser_exe_path

            browser = p.chromium.launch(**launch_options)
            page = browser.new_page(
                viewport={'width': 1200, 'height': 1600},
                device_scale_factor=1.0
            )

            html_file_path = Path(html_path).absolute()
            html_file_url = html_file_path.as_uri()
            page.goto(html_file_url, wait_until="networkidle")

            page.wait_for_load_state("load")
            page.wait_for_timeout(1000)

            page.evaluate("""
                () => {
                    const images = document.querySelectorAll('img');
                    return Promise.all(Array.from(images).map(img => {
                        if (img.complete) return Promise.resolve();
                        return new Promise(resolve => {
                            img.onload = resolve;
                            img.onerror = resolve;
                        });
                    }));
                }
            """)

            page.pdf(
                path=output_path,
                format="A4",
                print_background=True,
                display_header_footer=False,
                outline=True,
                tagged=True,
                scale=1.0,
                prefer_css_page_size=False,
                margin={
                    "top": "20mm",
                    "right": "15mm",
                    "bottom": "20mm",
                    "left": "15mm"
                }
            )
            
            browser.close()
    
    def generate_pdf(self, html_path: str, output_path: Optional[str] = None) -> str:
        html_file = Path(html_path)
        if not html_file.exists():
            raise FileNotFoundError(f"HTML文件不存在: {html_path}")
        
        if output_path is None:
            output_dir = Path("output/pdf")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            output_filename = html_file.stem + ".pdf"
            output_path = output_dir / output_filename
        else:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"📄 生成PDF: {html_file.name} -> {output_path.name}")

        if output_path.exists() and is_file_locked(output_path):
            error_msg = (
                f"目标PDF文件被占用，无法覆盖: {output_path}\n"
                f"请关闭以下可能打开该文件的程序后重试:\n"
                f"  - Adobe Acrobat / Reader\n"
                f"  - Microsoft Edge\n"
                f"  - Google Chrome\n"
                f"  - 其他PDF阅读器"
            )
            logger.error(f"❌ {error_msg}")
            raise PermissionError(error_msg)

        if output_path.exists():
            logger.info(f"🔄 删除旧文件: {output_path.name}")
            if not remove_file_with_retry(output_path):
                error_msg = f"无法删除旧PDF文件: {output_path}"
                logger.error(f"❌ {error_msg}")
                raise PermissionError(error_msg)

        return _generate_with_retry(self._generate_pdf, html_file, output_path, "PDF")
    
    def batch_generate_pdfs(self, html_dir: str = "output/html", show_progress: bool = True) -> list:
        html_path = Path(html_dir)
        if not html_path.exists():
            logger.warning(f"HTML目录不存在: {html_dir}")
            return []
        
        html_files = sorted(html_path.glob("*.html"))
        
        if not html_files:
            logger.warning(f"未找到HTML文件: {html_dir}")
            return []
        
        logger.info(f"🔄 批量生成PDF: 共 {len(html_files)} 个文件")
        
        pdf_files = []
        success_count = 0
        fail_count = 0
        
        if show_progress:
            from tqdm import tqdm
            pbar = tqdm(html_files, desc="📄 生成PDF")
        else:
            pbar = html_files
        
        for html_file in pbar:
            try:
                pdf_path = self.generate_pdf(str(html_file))
                pdf_files.append(pdf_path)
                success_count += 1
            except Exception as e:
                logger.error(f"❌ 生成失败: {html_file.name} - {e}")
                fail_count += 1
            
            if show_progress and hasattr(pbar, 'set_postfix'):
                pbar.set_postfix({
                    '成功': success_count,
                    '失败': fail_count
                })
        
        logger.info(f"✅ PDF批量生成完成: 成功 {success_count}/{len(html_files)}, 失败 {fail_count}/{len(html_files)}")
        
        return pdf_files

class DOCXGenerator:
    
    def __init__(self):
        self._check_dependencies()
    
    def _check_dependencies(self):
        try:
            import pypandoc

            pandoc_search_paths = [
                _project_base / "_internal" / "pandoc" / "pandoc.exe",
                _project_base / "pandoc" / "pandoc.exe",
            ]

            for pandoc_path in pandoc_search_paths:
                if pandoc_path.exists():
                    os.environ['PYPANDOC_PANDOC'] = str(pandoc_path.absolute())
                    try:
                        version = pypandoc.get_pandoc_version()
                        logger.info(f"✅ 使用项目 Pandoc {version}: {pandoc_path.relative_to(_project_base)}")
                    except:
                        logger.info(f"✅ 使用项目 Pandoc: {pandoc_path.relative_to(_project_base)}")
                    return

            if not _is_frozen:
                try:
                    version = pypandoc.get_pandoc_version()
                    logger.info(f"✅ 使用系统安装的 Pandoc {version}")
                    return
                except OSError:
                    pass
            
            raise RuntimeError(
                f"❌ Pandoc未安装\n\n"
                f"{'='*60}\n"
                f"📥 解决方案：运行依赖管理脚本\n"
                f"{'='*60}\n\n"
                f"请在项目目录运行:\n"
                f"  python setup_dependencies.py\n\n"
                f"此脚本将自动:\n"
                f"  ✓ 下载Pandoc到项目目录\n"
                f"  ✓ 配置所有必需的依赖\n"
                f"  ✓ 确保开发和打包环境正常\n\n"
                f"{'='*60}\n"
                f"💡 或者手动安装:\n"
                f"{'='*60}\n"
                f"  1. 访问: https://pandoc.org/installing.html\n"
                f"  2. 下载Windows版本\n"
                f"  3. 解压到 pandoc/ 目录"
            )
                
        except ImportError as e:
            raise RuntimeError(
                f"❌ pypandoc未安装\n"
                f"请运行: pip install pypandoc\n"
                f"错误: {e}"
            )
    
    def _set_formal_fonts(self, docx_path):
        try:
            from docx import Document
            doc = Document(str(docx_path))
            for para in doc.paragraphs:
                for run in para.runs:
                    run.font.name = 'Times New Roman'
            doc.save(str(docx_path))
        except Exception as e:
            logger.warning(f"设置字体失败（非致命）: {e}")

    def _preprocess_html(self, html_content: str) -> str:
        html_content = re.sub(
            r'<aside[^>]*>.*?</aside>',
            '',
            html_content,
            flags=re.DOTALL
        )
        
        html_content = re.sub(
            r'<button[^>]*class="menu-toggle"[^>]*>.*?</button>',
            '',
            html_content,
            flags=re.DOTALL
        )
        
        html_content = re.sub(
            r'<script[^>]*>.*?</script>',
            '',
            html_content,
            flags=re.DOTALL
        )
        
        html_content = html_content.replace(
            'class="main-content"',
            'class="main-content" style="margin-left: 0;"'
        )
        
        return html_content
    
    def _generate_docx(self, html_path: str, output_path: str):
        import pypandoc
        
        with open(html_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        html_content = self._preprocess_html(html_content)
        
        extra_args = [
            '--standalone',
            '--toc',
            '--toc-depth=3',
        ]
        
        if Path('reference.docx').exists():
            extra_args.append('--reference-doc=reference.docx')
        
        pypandoc.convert_text(
            html_content,
            'docx',
            format='html',
            outputfile=output_path,
            extra_args=extra_args
        )
        
        self._set_formal_fonts(output_path)
    
    def generate_docx(self, html_path: str, output_path: Optional[str] = None) -> str:
        html_file = Path(html_path)
        if not html_file.exists():
            raise FileNotFoundError(f"HTML文件不存在: {html_path}")
        
        if output_path is None:
            output_dir = Path("output/docx")
            output_dir.mkdir(parents=True, exist_ok=True)
            
            output_filename = html_file.stem + ".docx"
            output_path = output_dir / output_filename
        else:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"📝 生成DOCX: {html_file.name} -> {output_path.name}")

        if output_path.exists() and is_file_locked(output_path):
            error_msg = (
                f"目标DOCX文件被占用，无法覆盖: {output_path}\n"
                f"请关闭以下可能打开该文件的程序后重试:\n"
                f"  - Microsoft Word\n"
                f"  - WPS Office\n"
                f"  - LibreOffice Writer\n"
                f"  - 其他文档编辑器"
            )
            logger.error(f"❌ {error_msg}")
            raise PermissionError(error_msg)

        if output_path.exists():
            logger.info(f"🔄 删除旧文件: {output_path.name}")
            if not remove_file_with_retry(output_path):
                error_msg = f"无法删除旧DOCX文件: {output_path}"
                logger.error(f"❌ {error_msg}")
                raise PermissionError(error_msg)

        return _generate_with_retry(self._generate_docx, html_file, output_path, "DOCX")
    
    def batch_generate_docx(self, html_dir: str = "output/html", show_progress: bool = True) -> list:
        html_path = Path(html_dir)
        if not html_path.exists():
            logger.warning(f"HTML目录不存在: {html_dir}")
            return []
        
        html_files = sorted(html_path.glob("*.html"))
        
        if not html_files:
            logger.warning(f"未找到HTML文件: {html_dir}")
            return []
        
        logger.info(f"🔄 批量生成DOCX: 共 {len(html_files)} 个文件")
        
        docx_files = []
        success_count = 0
        fail_count = 0
        
        if show_progress:
            from tqdm import tqdm
            pbar = tqdm(html_files, desc="📝 生成DOCX")
        else:
            pbar = html_files
        
        for html_file in pbar:
            try:
                docx_path = self.generate_docx(str(html_file))
                docx_files.append(docx_path)
                success_count += 1
            except Exception as e:
                logger.error(f"❌ 生成失败: {html_file.name} - {e}")
                fail_count += 1
            
            if show_progress and hasattr(pbar, 'set_postfix'):
                pbar.set_postfix({
                    '成功': success_count,
                    '失败': fail_count
                })
        
        logger.info(f"✅ DOCX批量生成完成: 成功 {success_count}/{len(html_files)}, 失败 {fail_count}/{len(html_files)}")

        return docx_files

class ExcelGenerator:

    def __init__(self, output_dir: str = "output/excel"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _clean_xml_incompatible_chars(text):
        if not isinstance(text, str):
            return text
        illegal_xml_chars = [
            (0x00, 0x08),
            (0x0B, 0x0C),
            (0x0E, 0x1F),
            (0x7F, 0x84),
            (0x86, 0x9F),
        ]
        pattern = '|'.join([
            f'[\\x{start:02x}-\\x{end:02x}]'
            for start, end in illegal_xml_chars
        ])
        return re.sub(pattern, '', text)

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df_cleaned = df.copy()
        for col in df_cleaned.columns:
            if df_cleaned[col].dtype == 'object':
                df_cleaned[col] = df_cleaned[col].apply(
                    self._clean_xml_incompatible_chars
                )
        return df_cleaned

    def generate_excel(self, df: pd.DataFrame, pdf_name: str, is_translated: bool = False) -> str:
        if is_translated:
            filename = f"(译文) {pdf_name}.xlsx"
        else:
            filename = f"{pdf_name}.xlsx"
        return self.generate_excel_with_path(df, str(self.output_dir / filename), is_translated)

    def generate_excel_with_path(self, df: pd.DataFrame, output_path: str, is_translated: bool = False) -> str:
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        print(f"[Excel] 生成Excel: {output_file.name}")
        df_cleaned = self._clean_dataframe(df)
        df_cleaned.to_excel(output_file, index=False, engine='openpyxl')
        self._format_excel(output_file)
        print(f"[Excel] Excel已保存: {output_file}\n")
        return str(output_file)

    def _format_excel(self, excel_path: str):
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            column_widths = {'A': 40, 'B': 30, 'C': 25, 'D': 80}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for row in ws.iter_rows(min_row=2):
                ws.row_dimensions[row[0].row].height = None
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.freeze_panes = "A2"
            wb.save(excel_path)
        except Exception as e:
            print(f"[WARNING] Excel格式化失败: {str(e)}")
